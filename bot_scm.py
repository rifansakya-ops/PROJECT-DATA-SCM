import time
import requests
import json
import io
import os
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

def jalankan_bot():
    print("=== [LOG START] Memulai Operasi Bot SCM ===")
    
    # 1. Konfigurasi Browser Headless
    options = uc.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    
    print("[1/6] Menyiapkan Browser Siluman (Versi 145)...")
    # Mengunci ke versi 145 untuk menghindari SessionNotCreatedException
    driver = uc.Chrome(options=options, version_main=145)
    
    try:
        # 2. Buka Halaman Login
        print("[2/6] Membuka Halaman Login: https://scm.nusadaya.net/login")
        driver.get("https://scm.nusadaya.net/login")
        wait = WebDriverWait(driver, 25)
        
        # 3. Ambil Kredensial dari GitHub Secrets
        email_rahasia = os.environ.get('EMAIL_SCM')
        pass_rahasia = os.environ.get('PASS_SCM')

        print("[3/6] Mengisi Form Login...")
        email_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='text' or @placeholder='Email atau NIP']")))
        email_input.send_keys(email_rahasia)
        
        pass_input = driver.find_element(By.XPATH, "//input[@type='password']")
        pass_input.send_keys(pass_rahasia)
        
        print("-> Mengetuk tombol Log in...")
        driver.find_element(By.XPATH, "//button[contains(text(), 'Log in')]").click()
        
        print("-> Menunggu dashboard (15 detik)...")
        time.sleep(15)
        
        # 4. Ambil Cookies & Download Excel
        print("[4/6] Menembak URL Export Langsung...")
        cookies = driver.get_cookies()
        session_cookies = {c['name']: c['value'] for c in cookies}
        
        export_url = "https://scm.nusadaya.net/monitoring-kontrak-rinci/export?khs=all&bidang=all&tahun=all&stage="
        response_dl = requests.get(export_url, cookies=session_cookies)
        
        if response_dl.status_code == 200:
            print(f"-> Download Sukses! Ukuran: {len(response_dl.content)} bytes")
        else:
            print(f"(!) Gagal download. Status Code: {response_dl.status_code}")
            return

        # 5. Membongkar Excel & Ekstrak Link
        print("[5/6] Membongkar isi Excel & Mengekstrak Hyperlink...")
        wb = load_workbook(filename=io.BytesIO(response_dl.content), data_only=False)
        ws = wb.active
        
        data_rows = []
        for row in ws.iter_rows(min_row=2):
            current_row = []
            for cell in row:
                if cell.hyperlink:
                    label = str(cell.value) if cell.value is not None else "Lihat File"
                    url = cell.hyperlink.target
                    current_row.append(f"{label} {url}")
                else:
                    current_row.append(cell.value if cell.value is not None else "")
            data_rows.append(current_row)

        print(f"-> Total data: {len(data_rows)} baris siap kirim.")

        # 6. Kirim ke Google Sheets
        print("[6/6] Mengirim data ke Google Sheets...")
        gas_url = "https://script.google.com/macros/s/AKfycbwdAuwnHlrwH8SAlpkeBp6fUcCVNqINvNuNQpeeLRUzfuZNq3NmUXMwr-_6TXgKkJhOww/exec"
        
        payload = json.dumps({"rows": data_rows})
        res_gas = requests.post(gas_url, data=payload, headers={'Content-Type': 'application/json'})
        
        print(f"=== [HASIL AKHIR] Respon GAS: {res_gas.text} ===")

    except Exception as e:
        print(f"!!! [ERROR] Terjadi kendala teknis: {str(e)}")
    finally:
        print("=== [LOG END] Menutup Browser & Selesai ===")
        driver.quit()

if __name__ == "__main__":
    jalankan_bot()
